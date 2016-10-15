import question_module as s
import re
import traceback
from openpyxl import load_workbook
import csv

# Close and move onto the next stop
#output_csv.close()
print('-' * 70)

print("\nWelcome to question or not. Please input the sentence you wish to analyze.")

sentence = ''
while not (sentence == 'exit'):
    print('\n')
    print('-' * 70)
    sentence = input('Neg = Not a question, Pos = Question\nTo finish the program please type in exit.\nPleae input the sentence you wish to analyze : ')

    try:
        value = s.sentiment(str(sentence))
        print('-' * 70)
        print('Classifier with the stop words predicted : ',value[0])
        print('confidence of the above prediction : ',value[1])
        print('Classifier without the stop words predicted : ',value[2])
        print('confidence of the above prediction : ',value[3])
    except:
        traceback.print_exc()
        print('\nPlease do not include any special char.')
        print('-' * 70)
        pass

print('Good bye! Thank you.')
print('Question or Not by Jae Duk Seo 2016')
print('Copy right to Jae Duk Seo')

#----------COMMENTS THAT I DONT KNOW WHY THIS IS HERE-------------
# Reading the existing xlsx book
#workbook = load_workbook(filename = 'jurassic.xlsx',use_iterators=True)
#first_sheet = workbook.get_sheet_names()[0] #Name of the work sheet
#worksheet = workbook.get_sheet_by_name(first_sheet) # Read the first worksheet
#number_of_col = worksheet.max_row

# Open the csv file and write the basic info
#output_csv = open('out.csv','w')
#output_csv.write('Title,With,With conf,No,No conf,Text,With,With conf,No,No conf\n')

#for i in range(2,number_of_col):
    # cur_title = str(worksheet['G' + str(i)].value)  # Title value
    # cur_text =  str(worksheet['H' + str(i)].value)  # Text value
    #
    # if cur_title != 'None':
    #     cur_title_value = s.sentiment(cur_title)
    #     if str(cur_title_value[0]) == 'neg':
    #         temp = 'Not a question'
    #     elif str(cur_title_value[0]) == 'pos':
    #         temp = 'Question'
    #     else:
    #         temp = "Neutrual"
    #
    #     if str(cur_title[2]) == 'neg':
    #         temp2 = 'Not a question'
    #     elif str(cur_title_value[2]) == 'pos':
    #         temp2 = 'Question'
    #     else:
    #         temp2 = "Neutrual"
    #
    #     output_csv.write(str(cur_title).replace(',',' ') + ',' +
    #                      str(temp) + ',' +
    #                      str(cur_title_value[1]) + ',' +
    #                      str(temp2) + ',' +
    #                      str(cur_title_value[3]))
    #
    # if cur_text != "None":
    #     cur_text_value = s.sentiment(cur_text)
    #     if str(cur_text_value[0]) == 'neg':
    #         temp = 'Not a question'
    #     elif str(cur_text_value[0]) == 'pos':
    #         temp = 'Question'
    #     else:
    #         temp = "Neutrual"
    #
    #     if str(cur_text_value[2]) == 'neg':
    #         temp2 = 'Not a question'
    #     elif str(cur_text_value[2]) == 'pos':
    #         temp2 = 'Question'
    #     else:
    #         temp2 = "Neutrual"
    #
    #     output_csv.write(',,,,,' +str(cur_text).replace(',',' ') + ',' +
    #                               str(temp) + ',' +
    #                               str(cur_text_value[1]) + ',' +
    #                               str(temp2) + ',' +
    #                               str(cur_text_value[3]) + '\n')
